from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "订单编号", "订单信息", "项目编号", "业务", "数量", "订单类", "BOM", "机芯", "FORMAT",
    "EMMC", "需求时间", "发布时间", "状态", "SPM", "版本", "ID", "备注",
]
# Order number is the primary key. Within the rolling 8-month RDM window,
# movement/engine code can be corrected by RDM, so only quantity stays locked.
LOCKED_FIELDS = ("数量",)


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def quantity(value) -> Decimal:
    raw = text(value).replace(",", "") or "0"
    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"数量无法解析：{value}") from exc


def locked_equal(field: str, old, new) -> bool:
    if field == "数量":
        return quantity(old) == quantity(new)
    return text(old).upper() == text(new).upper()


def atomic_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, path)


def write_conflict_workbook(path: Path, report: dict) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据锁定冲突"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "Fire TV 出货量数据锁定冲突"
    sheet["A1"].fill = PatternFill("solid", fgColor="C00000")
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    sheet.row_dimensions[1].height = 28
    sheet.append(["截止日期", report["cutoffDate"], "状态", "已阻断发布", "冲突数", report["conflictCount"], "原始数据是否更新", "否"])
    sheet.append(["处理原则", "确认锁定字段后重跑", "旧页面", "继续保留", "RDM快照", "已保存", "通知", "微信告警"])
    sheet.append([])
    headers = ["状态", "订单编号", "字段", "历史值", "RDM当前值", "检测时间", "处理决定", "备注"]
    sheet.append(headers)
    for cell in sheet[5]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for item in report["conflicts"]:
        sheet.append(["待确认", item["orderNo"], item.get("field", ""), item.get("previousValue", ""), item.get("rdmValue", ""), report["generatedAt"], "", item["message"]])
    status_validation = DataValidation(type="list", formula1='"待确认,已确认,已作废"')
    decision_validation = DataValidation(type="list", formula1='"采用RDM值,保留历史值,修正数据后重跑"')
    sheet.add_data_validation(status_validation)
    sheet.add_data_validation(decision_validation)
    status_validation.add("A6:A100")
    decision_validation.add("G6:G100")
    sheet.freeze_panes = "A6"
    widths = [12, 18, 14, 22, 22, 24, 20, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw", required=True)
    parser.add_argument("--snapshot", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--backup-dir", required=True)
    parser.add_argument("--cutoff", required=True)
    parser.add_argument("--conflict-xlsx", default="")
    args = parser.parse_args()

    raw_path = Path(args.raw)
    snapshot_path = Path(args.snapshot)
    report_path = Path(args.report)
    backup_dir = Path(args.backup_dir)

    with snapshot_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != HEADERS:
            raise ValueError(f"RDM列名不完整。实际：{reader.fieldnames}")
        downloaded = [{header: text(row.get(header)) for header in HEADERS} for row in reader]

    if not downloaded:
        raise ValueError("RDM下载结果为0行")
    if any(not row["订单编号"] for row in downloaded):
        raise ValueError("RDM下载结果存在空订单编号")
    for row in downloaded:
        quantity(row["数量"])

    snapshot_groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in downloaded:
        snapshot_groups[row["订单编号"]].append(row)

    snapshot_conflicts = []
    snapshot_rows: dict[str, dict[str, str]] = {}
    exact_snapshot_duplicates = 0
    for order_no, rows in snapshot_groups.items():
        fingerprints = {tuple(row[header] for header in HEADERS) for row in rows}
        if len(fingerprints) > 1:
            snapshot_conflicts.append({
                "orderNo": order_no,
                "type": "snapshot_duplicate_conflict",
                "message": "同一订单编号在本次RDM下载中出现不同内容",
            })
        else:
            exact_snapshot_duplicates += len(rows) - 1
            snapshot_rows[order_no] = rows[0]

    workbook = load_workbook(raw_path)
    sheet = workbook.active
    raw_indices: dict[str, list[int]] = defaultdict(list)
    for row_index in range(2, sheet.max_row + 1):
        order_no = text(sheet.cell(row_index, 1).value)
        if order_no:
            raw_indices[order_no].append(row_index)

    locked_conflicts = []
    for order_no, new_row in snapshot_rows.items():
        for row_index in raw_indices.get(order_no, []):
            for field in LOCKED_FIELDS:
                column_index = HEADERS.index(field) + 1
                old_value = sheet.cell(row_index, column_index).value
                if not locked_equal(field, old_value, new_row[field]):
                    locked_conflicts.append({
                        "orderNo": order_no,
                        "field": field,
                        "previousValue": text(old_value),
                        "rdmValue": new_row[field],
                        "sourceRow": row_index,
                        "type": "locked_field_changed",
                        "message": "锁定字段与历史原始数据不一致",
                    })

    conflicts = snapshot_conflicts + locked_conflicts
    base_report = {
        "schemaVersion": 1,
        "cutoffDate": args.cutoff,
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "snapshotFile": str(snapshot_path),
        "rawFile": str(raw_path),
        "downloadedRows": len(downloaded),
        "downloadedUniqueOrders": len(snapshot_rows),
        "exactSnapshotDuplicateRowsRemoved": exact_snapshot_duplicates,
        "existingOrdersMatched": sum(1 for order_no in snapshot_rows if order_no in raw_indices),
        "newOrders": sum(1 for order_no in snapshot_rows if order_no not in raw_indices),
        "conflictCount": len(conflicts),
        "conflicts": conflicts,
    }

    if conflicts:
        base_report["status"] = "blocked"
        base_report["rawUpdated"] = False
        atomic_json(report_path, base_report)
        if args.conflict_xlsx:
            write_conflict_workbook(Path(args.conflict_xlsx), base_report)
        print(json.dumps(base_report, ensure_ascii=True))
        return 2

    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"Fire TV quantity - raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(raw_path, backup_path)

    duplicate_rows_to_delete = []
    updated_orders = 0
    added_orders = 0
    for order_no, new_row in snapshot_rows.items():
        row_numbers = raw_indices.get(order_no, [])
        if row_numbers:
            target_row = row_numbers[0]
            duplicate_rows_to_delete.extend(row_numbers[1:])
            updated_orders += 1
        else:
            target_row = sheet.max_row + 1
            added_orders += 1
        for column_index, header in enumerate(HEADERS, start=1):
            sheet.cell(target_row, column_index).value = new_row[header]

    for row_index in sorted(set(duplicate_rows_to_delete), reverse=True):
        sheet.delete_rows(row_index, 1)

    temp_path = raw_path.with_suffix(".xlsx.tmp")
    workbook.save(temp_path)
    os.replace(temp_path, raw_path)

    base_report.update({
        "status": "success",
        "rawUpdated": True,
        "backupFile": str(backup_path),
        "updatedOrders": updated_orders,
        "addedOrders": added_orders,
        "existingDuplicateRowsRemoved": len(set(duplicate_rows_to_delete)),
    })
    atomic_json(report_path, base_report)
    print(json.dumps(base_report, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise
