from __future__ import annotations

import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


DASHBOARD_DIR = Path(__file__).resolve().parent
WORKSPACE = DASHBOARD_DIR.parents[1]
SHARED_ROOT = Path(os.environ.get("FIRE_TV_SHARED_ROOT", ".")).resolve()
LOCAL_RULES_DIR = WORKSPACE / "outputs" / "fire_tv_rules"


def resolve_path(env_name: str, candidates: list[Path], label: str) -> Path:
    override = os.environ.get(env_name)
    if override:
        path = Path(override)
        if not path.exists():
            raise FileNotFoundError(f"{label}不存在：{path}")
        return path
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(f"未找到{label}：{'；'.join(str(path) for path in candidates)}")


RULE_PATH = resolve_path(
    "FIRE_TV_RULE_PATH",
    [SHARED_ROOT / "Fire TV quantity - rule.xlsx"],
    "机芯口径规则表",
)
CALCULATION_RULE_PATH = resolve_path(
    "FIRE_TV_CALCULATION_RULE_PATH",
    [
        SHARED_ROOT / "Fire TV quantity - calculation-rules.json",
        LOCAL_RULES_DIR / "Fire TV quantity - calculation-rules.json",
    ],
    "统计计算规则",
)
CALCULATION_RULES = json.loads(CALCULATION_RULE_PATH.read_text(encoding="utf-8"))
RAW_PATH = resolve_path(
    "FIRE_TV_RAW_PATH",
    [SHARED_ROOT / "Fire TV quantity - raw.xlsx", SHARED_ROOT / "20260623.xlsx"],
    "原始主数据",
)
PATCH_PATH = None
if os.environ.get("FIRE_TV_DISABLE_PATCH", "").strip() != "1" and CALCULATION_RULES.get("sourceFiles", {}).get("patchActive", True):
    PATCH_PATH = resolve_path(
        "FIRE_TV_PATCH_PATH",
        [SHARED_ROOT / "Fire TV quantity - patch.xlsx", SHARED_ROOT / "data" / "赖锦灵补丁.xlsx"],
        "补丁数据",
    )

VALID_STATUSES = set(CALCULATION_RULES["validStatuses"])
NO_SOFTWARE = "不需软件"
NO_SOFTWARE_RULES = {rule["id"]: rule for rule in CALCULATION_RULES["noSoftwareRules"]}
PR_TRIAL = NO_SOFTWARE_RULES["EX-04"]["orderTypeEquals"]
G_BUSINESSES = set(NO_SOFTWARE_RULES["EX-01"]["businessEqualsAny"])
G_BOM_REGEX = NO_SOFTWARE_RULES["EX-01"]["bomRegex"]
B_BUSINESSES = set(NO_SOFTWARE_RULES["EX-03"]["businessEqualsAny"])
B_BOM_REGEX = NO_SOFTWARE_RULES["EX-03"]["bomRegex"]
M_BUSINESSES = set(NO_SOFTWARE_RULES["EX-02"]["businessEqualsAny"])
M_BOM_MARKER = NO_SOFTWARE_RULES["EX-02"]["bomContains"].upper()
M_LINKED_RULE = NO_SOFTWARE_RULES["EX-02"]["linkedVersionRule"]
START_YEAR = int(CALCULATION_RULES["reportingRange"]["startYear"])
CONFIGURED_END_YEAR = int(CALCULATION_RULES["reportingRange"]["endYear"])
END_YEAR = max(CONFIGURED_END_YEAR, datetime.now().year) if CALCULATION_RULES["reportingRange"].get("endYearMode") == "current_year" else CONFIGURED_END_YEAR
YEARS = list(range(START_YEAR, END_YEAR + 1))

REASON_LABELS = {
    "included": "计入统计",
    "invalid_status": "状态不在统计范围",
    "pr_trial": "不需软件 + PR试产",
    "g_us_mx": "G客户 US/MX BOM减免",
    "b_us": "B客户 US BOM减免",
    "m_in44": "M客户 IN4.4 BOM减免",
    "m_linked": "M客户关联版本减免",
    "duplicate_exact": "主数据与补丁完全重复",
    "duplicate_conflict": "同订单号跨来源明细冲突",
    "unassigned": "重复机芯未能归属",
    "missing_effective_date": "发布时间和需求时间均为空",
    "date_outside_range": "统计时间超出报表范围",
}
BLOCKING_REASON_TYPES = {
    "unassigned": "duplicate_routing_unassigned",
    "duplicate_conflict": "duplicate_conflict",
    "missing_effective_date": "missing_effective_date",
    "date_outside_range": "date_outside_reporting_range",
}


def clean(value: Any) -> str:
    return str(value or "").strip()


def numeric(value: Any) -> int | float:
    number = float(value or 0)
    return int(number) if number.is_integer() else number


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def iso_date(value: Any) -> str | None:
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else None


def split_order_number(value: str) -> tuple[str, int | None]:
    match = re.match(r"^(.*?)(\d+)$", value)
    if not match:
        return value, None
    return match.group(1), int(match.group(2))


def normalized_in_version(bom: str) -> str:
    return re.sub(r"IN4\.[234]", "IN4.X", bom.upper())


def unique(values: list[str]) -> list[str]:
    return list(dict.fromkeys(value for value in values if value))


def split_engine_codes(value: Any) -> list[str]:
    return unique([part.strip() for part in re.split(r"[\n,，;；、]+", clean(value)) if part.strip()])


def merged_value_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, column)] = value
    return values


def read_rule_groups() -> tuple[list[dict[str, Any]], dict[str, list[str]], str, dict[str, Any]]:
    workbook = openpyxl.load_workbook(RULE_PATH, data_only=True)
    sheet = workbook.worksheets[0]
    if sheet.max_column < 14 or clean(sheet.cell(1, 10).value) != "口径ID":
        raise ValueError("规则表尚未升级稳定口径ID，请先运行统一结构安装。")
    merged_values = merged_value_map(sheet)

    def value(row: int, column: int) -> Any:
        direct = sheet.cell(row, column).value
        return direct if direct is not None else merged_values.get((row, column))

    groups_by_id: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    missing_id_rows: list[int] = []
    for row in range(3, sheet.max_row + 1):
        engines = split_engine_codes(sheet.cell(row, 6).value)
        if not engines:
            continue
        project_id = clean(sheet.cell(row, 10).value)
        if not project_id:
            missing_id_rows.append(row)
            continue
        if not re.fullmatch(r"FTV-\d{4,}", project_id):
            raise ValueError(f"规则表第{row}行口径ID格式错误：{project_id}")
        if project_id not in groups_by_id:
            groups_by_id[project_id] = {
                "id": project_id,
                "sourceRow": row,
                "customer": clean(value(row, 1)),
                "category": clean(value(row, 2)),
                "project": clean(value(row, 3)),
                "chip": clean(value(row, 4)),
                "region": clean(value(row, 5)),
                "engines": [],
                "bomRegions": [],
                "cooperationModes": [],
                "osVersions": [],
                "effectiveDate": iso_date(sheet.cell(row, 11).value),
                "retroactive": clean(sheet.cell(row, 12).value) == "是",
                "ruleVersion": clean(sheet.cell(row, 13).value),
                "changeReason": clean(sheet.cell(row, 14).value),
            }
            order.append(project_id)
        group = groups_by_id[project_id]
        group["engines"] = unique([*group["engines"], *engines])
        group["bomRegions"] = unique([*group["bomRegions"], clean(value(row, 7))])
        group["cooperationModes"] = unique([*group["cooperationModes"], clean(value(row, 8))])
        group["osVersions"] = unique([*group["osVersions"], clean(value(row, 9))])

    if missing_id_rows:
        raise ValueError(f"规则表以下机芯行缺少口径ID：{missing_id_rows}")
    groups = [groups_by_id[project_id] for project_id in order]
    if not groups:
        raise ValueError("规则表没有可用的项目口径。")

    engine_to_groups: dict[str, list[str]] = defaultdict(list)
    for group in groups:
        for engine in group["engines"]:
            engine_to_groups[engine].append(group["id"])
        group["total"] = 0
        group["annual"] = {str(year): 0 for year in YEARS}
        group["quarterly"] = {f"{year}Q{quarter}": 0 for year in YEARS for quarter in range(1, 5)}

    coverage = {
        "rulePath": str(SHARED_ROOT / "Fire TV quantity - rule.xlsx"),
        "ruleSheet": sheet.title,
        "ruleProjectCount": len(groups),
        "ruleEngineCount": len(engine_to_groups),
        "duplicateEngineCodes": sorted(engine for engine, ids in engine_to_groups.items() if len(ids) > 1),
        "missingProjectIdRows": [],
    }
    return groups, dict(engine_to_groups), sheet.title, coverage


def group_lookup(groups: list[dict[str, Any]], customer: str, project: str) -> str | None:
    for group in groups:
        if group["customer"] == customer and group["project"] == project:
            return group["id"]
    return None


def assign_group(engine: str, business: str, groups: list[dict[str, Any]], engine_to_groups: dict[str, list[str]]) -> str | None:
    candidates = engine_to_groups[engine]
    if len(candidates) == 1:
        return candidates[0]
    for routing in CALCULATION_RULES.get("duplicateEngineRouting", []):
        if routing["engine"] != engine:
            continue
        for route in routing["routes"]:
            if any(token in business for token in route["businessContainsAny"]):
                target = group_lookup(groups, route["customer"], route["project"])
                if target in candidates:
                    return target
    return None


def exclusion_reason(status: str, order_type: str, business: str, bom: str) -> str | None:
    if status not in VALID_STATUSES:
        return "invalid_status"
    if status == NO_SOFTWARE and order_type == PR_TRIAL:
        return "pr_trial"
    if status == NO_SOFTWARE and business in G_BUSINESSES and re.search(G_BOM_REGEX, bom, re.IGNORECASE):
        return "g_us_mx"
    if status == NO_SOFTWARE and business in B_BUSINESSES and re.search(B_BOM_REGEX, bom, re.IGNORECASE):
        return "b_us"
    if status == NO_SOFTWARE and business in M_BUSINESSES and M_BOM_MARKER in bom.upper():
        return "m_in44"
    return None


def fingerprint_value(value: Any) -> str:
    parsed = parse_date(value)
    if parsed:
        return parsed.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def row_fingerprint(row: tuple[Any, ...]) -> str:
    return "\u001f".join(fingerprint_value(value) for value in row)


def read_audit_records(groups: list[dict[str, Any]], engine_to_groups: dict[str, list[str]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    records: list[dict[str, Any]] = []
    group_by_id = {group["id"]: group for group in groups}
    primary_fingerprints: set[str] = set()
    primary_orders: dict[str, set[str]] = defaultdict(set)
    outside_scope_rows = 0
    outside_scope_engines: Counter[str] = Counter()

    source_paths = [("主数据", RAW_PATH)]
    if PATCH_PATH is not None:
        source_paths.append(("补丁", PATCH_PATH))
    for source_index, (source_name, raw_path) in enumerate(source_paths):
        workbook = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
        sheet = workbook.active
        for source_row, raw_row in enumerate(sheet.iter_rows(min_row=2, max_col=17, values_only=True), start=2):
            row = tuple(raw_row)
            engine = clean(row[7])
            if engine not in engine_to_groups:
                outside_scope_rows += 1
                if engine:
                    outside_scope_engines[engine] += 1
                continue

            fingerprint = row_fingerprint(row)
            order_no = clean(row[0])
            business = clean(row[3])
            group_id = assign_group(engine, business, groups, engine_to_groups)
            status = clean(row[12])
            order_type = clean(row[5])
            bom = clean(row[6])
            reason = "unassigned" if group_id is None else exclusion_reason(status, order_type, business, bom)

            if source_index == 0:
                primary_fingerprints.add(fingerprint)
                if order_no:
                    primary_orders[order_no].add(fingerprint)
            else:
                if fingerprint in primary_fingerprints:
                    reason = "duplicate_exact"
                elif order_no and order_no in primary_orders and fingerprint not in primary_orders[order_no]:
                    reason = "duplicate_conflict"

            effective = parse_date(row[11]) or parse_date(row[10])
            if reason is None and effective is None:
                reason = "missing_effective_date"
            if reason is None and effective and not START_YEAR <= effective.year <= END_YEAR:
                reason = "date_outside_range"
            year = effective.year if effective else None
            quarter = f"{year}Q{((effective.month - 1) // 3) + 1}" if effective else None
            group = group_by_id.get(group_id)
            records.append(
                {
                    "source": source_name,
                    "sourceFile": raw_path.name,
                    "sourceRow": source_row,
                    "groupId": group_id,
                    "customer": group["customer"] if group else "未归属",
                    "category": group["category"] if group else "",
                    "project": group["project"] if group else "",
                    "chip": group["chip"] if group else "",
                    "engine": engine,
                    "orderNo": order_no,
                    "business": business,
                    "quantity": numeric(row[4]),
                    "orderType": order_type,
                    "bom": bom,
                    "demandDate": iso_date(row[10]),
                    "releaseDate": iso_date(row[11]),
                    "effectiveDate": effective.isoformat() if effective else None,
                    "year": year,
                    "quarter": quarter,
                    "status": status,
                    "spm": clean(row[13]),
                    "reasonKey": reason or "included",
                    "included": reason is None,
                }
            )

    direct_m_exclusions = [record for record in records if record["reasonKey"] == "m_in44"]
    linked_versions = [version.upper() for version in M_LINKED_RULE["excludedVersions"]]
    for record in records:
        if not record["included"] or record["business"] not in M_BUSINESSES:
            continue
        if not any(version in record["bom"].upper() for version in linked_versions):
            continue
        prefix, suffix = split_order_number(record["orderNo"])
        for excluded in direct_m_exclusions:
            excluded_prefix, excluded_suffix = split_order_number(excluded["orderNo"])
            if (
                prefix == excluded_prefix
                and suffix is not None
                and excluded_suffix is not None
                and abs(suffix - excluded_suffix) <= int(M_LINKED_RULE["maxTrailingOrderNumberDifference"])
                and record["quantity"] == excluded["quantity"]
                and normalized_in_version(record["bom"]) == normalized_in_version(excluded["bom"])
            ):
                record["included"] = False
                record["reasonKey"] = "m_linked"
                break

    scope_summary = {
        "outsideScopeRows": outside_scope_rows,
        "outsideScopeUniqueEngines": len(outside_scope_engines),
        "topOutsideScopeEngines": [
            {"engine": engine, "rows": rows}
            for engine, rows in outside_scope_engines.most_common(20)
        ],
    }
    return records, scope_summary


def calculate_groups(groups: list[dict[str, Any]], records: list[dict[str, Any]]) -> None:
    group_by_id = {group["id"]: group for group in groups}
    for record in records:
        if not record["included"]:
            continue
        group = group_by_id[record["groupId"]]
        quantity = float(record["quantity"])
        group["total"] = numeric(float(group["total"]) + quantity)
        year_key = str(record["year"])
        group["annual"][year_key] = numeric(float(group["annual"][year_key]) + quantity)
        group["quarterly"][record["quarter"]] = numeric(float(group["quarterly"][record["quarter"]]) + quantity)


def summarize(records: list[dict[str, Any]]) -> dict[str, Any]:
    reason_totals: Counter[str] = Counter()
    reason_rows: Counter[str] = Counter()
    for record in records:
        reason_totals[record["reasonKey"]] += float(record["quantity"])
        reason_rows[record["reasonKey"]] += 1
    return {
        "matchedRows": len(records),
        "matchedQuantity": numeric(sum(float(record["quantity"]) for record in records)),
        "includedRows": reason_rows["included"],
        "includedQuantity": numeric(reason_totals["included"]),
        "excludedRows": len(records) - reason_rows["included"],
        "excludedQuantity": numeric(sum(value for key, value in reason_totals.items() if key != "included")),
        "byReason": [
            {"key": key, "label": REASON_LABELS[key], "rows": reason_rows[key], "quantity": numeric(reason_totals[key])}
            for key in REASON_LABELS
            if key != "included" and reason_rows[key]
        ],
    }


def build_exceptions(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    exceptions: list[dict[str, Any]] = []
    for record in records:
        reason = record["reasonKey"]
        if reason not in BLOCKING_REASON_TYPES and reason != "duplicate_exact":
            continue
        blocking = reason in BLOCKING_REASON_TYPES
        exceptions.append(
            {
                "type": BLOCKING_REASON_TYPES.get(reason, "duplicate_exact"),
                "severity": "阻断" if blocking else "提示",
                "source": record["source"],
                "sourceFile": record["sourceFile"],
                "sourceRow": record["sourceRow"],
                "orderNo": record["orderNo"],
                "engine": record["engine"],
                "business": record["business"],
                "quantity": record["quantity"],
                "message": REASON_LABELS[reason],
                "resolution": "",
            }
        )
    return exceptions


def build() -> dict[str, Any]:
    groups, engine_to_groups, rule_sheet, coverage = read_rule_groups()
    records, scope_summary = read_audit_records(groups, engine_to_groups)
    calculate_groups(groups, records)
    summary = summarize(records)
    exceptions = build_exceptions(records)
    blocking_count = sum(1 for item in exceptions if item["severity"] == "阻断")
    data_as_of_override = os.environ.get("FIRE_TV_DATA_AS_OF", "").strip()
    source_date_match = re.search(r"(20\d{6})", RAW_PATH.stem)
    if data_as_of_override:
        data_as_of = datetime.strptime(data_as_of_override, "%Y-%m-%d").date().isoformat()
    elif source_date_match:
        data_as_of = datetime.strptime(source_date_match.group(1), "%Y%m%d").date().isoformat()
    else:
        data_as_of = datetime.fromtimestamp(RAW_PATH.stat().st_mtime).date().isoformat()
    customers = sorted({group["customer"] for group in groups}, key=lambda value: (value != "区域", value))
    coverage.update(scope_summary)
    return {
        "schemaVersion": 2,
        "title": "Fire TV 出货量看板",
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "dataAsOf": data_as_of,
        "sourceFiles": [
            "Fire TV quantity - raw.xlsx",
            *(["Fire TV quantity - patch.xlsx"] if PATCH_PATH is not None else []),
            "Fire TV quantity - rule.xlsx",
            "Fire TV quantity - calculation-rules.json",
        ],
        "years": YEARS,
        "customers": customers,
        "reasonLabels": REASON_LABELS,
        "ruleCoverage": coverage,
        "calculationRuleSource": str(SHARED_ROOT / "Fire TV quantity - calculation-rules.json"),
        "calculationRuleVersion": CALCULATION_RULES["version"],
        "groups": groups,
        "exceptions": {
            "blockingCount": blocking_count,
            "items": exceptions,
        },
        "refreshBlocked": blocking_count > 0,
        "audit": {"summary": summary, "records": records},
    }


def atomic_write(path: Path, text: str) -> None:
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(text, encoding="utf-8")
    temp_path.replace(path)


def main() -> int:
    data = build()
    compact_json = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    atomic_write(DASHBOARD_DIR / "calculation-data.json", compact_json)
    if data["refreshBlocked"]:
        print(f"发现 {data['exceptions']['blockingCount']} 条阻断异常；已生成待核查数据，未覆盖现有页面数据。")
        return 2
    atomic_write(DASHBOARD_DIR / "dashboard-data.json", compact_json)
    atomic_write(DASHBOARD_DIR / "dashboard-data.js", f"window.FTV_DASHBOARD_DATA={compact_json};\n")
    summary = data["audit"]["summary"]
    print(
        f"完成：{len(data['groups'])} 个项目，"
        f"{summary['includedQuantity']:,} 台计入，"
        f"{summary['excludedQuantity']:,} 台排除。"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
